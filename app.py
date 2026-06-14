import os
import re
import json
import tempfile
from flask import Flask, request, abort
from linebot import LineBotApi, WebhookHandler
from linebot.exceptions import InvalidSignatureError
from linebot.models import MessageEvent, TextMessage, TextSendMessage
from google import genai as google_genai
import requests
import io
import openpyxl

app = Flask(__name__)

LINE_CHANNEL_ACCESS_TOKEN = os.environ["LINE_CHANNEL_ACCESS_TOKEN"]
LINE_CHANNEL_SECRET = os.environ["LINE_CHANNEL_SECRET"]
GEMINI_API_KEY = os.environ["GEMINI_API_KEY"]
GOOGLE_DRIVE_API_KEY = os.environ.get("GOOGLE_DRIVE_API_KEY", os.environ.get("GEMINI_API_KEY", ""))
SCHEDULE_FOLDER_ID = os.environ.get("SCHEDULE_FOLDER_ID", "1xoPDyT53_5OmrwiNnBmWcZ6t0JfgOCHj")
FALLBACK_SCHEDULE_FILE_ID = os.environ.get("SCHEDULE_FILE_ID", "1JuUl9oPPCbNA-KiGlWAqchatd06Ml4Zk")

line_bot_api = LineBotApi(LINE_CHANNEL_ACCESS_TOKEN)
handler = WebhookHandler(LINE_CHANNEL_SECRET)
gemini_client = google_genai.Client(api_key=GEMINI_API_KEY)


def get_current_week_file_id():
    """在 Google Drive 資料夾裡找最新的課表檔案"""
    url = "https://www.googleapis.com/drive/v3/files"
    params = {
        "q": f"'{SCHEDULE_FOLDER_ID}' in parents and trashed=false",
        "key": GOOGLE_DRIVE_API_KEY,
        "orderBy": "createdTime desc",
        "fields": "files(id,name)",
        "pageSize": 20,
    }
    try:
        resp = requests.get(url, params=params, timeout=10)
        if resp.status_code == 200:
            for f in resp.json().get("files", []):
                name = f["name"]
                if "課程表" in name or name.lower().endswith(".xlsx"):
                    return f["id"]
    except Exception:
        pass
    return FALLBACK_SCHEDULE_FILE_ID


def get_latest_schedule_text():
    """從 Google Drive 自動找最新課表 Excel 並轉成文字"""
    file_id = get_current_week_file_id()
    if not file_id:
        return None, None
    url = f"https://drive.google.com/uc?export=download&id={file_id}"
    resp = requests.get(url, timeout=30)
    if resp.status_code != 200:
        return None, None

    buf = io.BytesIO(resp.content)
    wb = openpyxl.load_workbook(buf, data_only=True)

    text_parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        text_parts.append(f"\n=== {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            row_text = "\t".join(str(c) if c is not None else "" for c in row)
            if row_text.strip():
                text_parts.append(row_text)

    return "\n".join(text_parts), "課表"


def search_materials(keywords):
    """用 Gemini 搜尋講義（回傳 Drive 搜尋連結）"""
    # 由於講義資料夾未公開，提供 Drive 搜尋連結讓老師自行點開
    query = "+".join(keywords)
    search_url = f"https://drive.google.com/drive/search?q={query}"
    return [(f"搜尋結果：{' '.join(keywords)}", search_url)]


def query_schedule_for_student(student_name, schedule_text):
    """用 Gemini 從課表文字找出學生的上課時間"""
    prompt = f"""以下是補習班課表資料：

{schedule_text}

請找出學生「{student_name}」的所有上課時間和科目，用清楚的格式列出。
如果找不到這個學生，請說「找不到 {student_name} 的課表資料，請確認姓名是否正確」。
只回覆課表資訊，不要加其他說明。"""
    response = gemini_client.models.generate_content(model="gemini-2.0-flash", contents=prompt)
    return response.text


def query_schedule_for_teacher(teacher_name, schedule_text):
    """用 Gemini 從課表文字找出老師這週的課"""
    prompt = f"""以下是補習班課表資料：

{schedule_text}

請找出老師「{teacher_name}」這週要教的所有學生、科目和上課時間，用清楚的格式列出。
如果找不到這個老師，請說「找不到 {teacher_name} 老師的課表資料，請確認姓名是否正確」。
只回覆課表資訊，不要加其他說明。"""
    response = gemini_client.models.generate_content(model="gemini-2.0-flash", contents=prompt)
    return response.text


def handle_message(text):
    text = text.strip()

    # 本週上課時間，王小明 / 本週上課時間，劉冠毅
    match_schedule = re.match(r"本週上課時間[，,]\s*(.+)", text)
    if match_schedule:
        name = match_schedule.group(1).strip()
        schedule_text, filename = get_latest_schedule_text()
        if not schedule_text:
            return "目前無法讀取課表，請稍後再試。"
        # 先嘗試當學生查，再嘗試當老師查，回傳兩者合併結果
        student_result = query_schedule_for_student(name, schedule_text)
        teacher_result = query_schedule_for_teacher(name, schedule_text)
        # 如果兩個都找不到
        if "找不到" in student_result and "找不到" in teacher_result:
            return f"找不到「{name}」的課表資料，請確認姓名是否正確。"
        parts = []
        if "找不到" not in student_result:
            parts.append(f"📅 {name} 的上課時間：\n{student_result}")
        if "找不到" not in teacher_result:
            parts.append(f"📋 {name} 老師本週的課：\n{teacher_result}")
        return "\n\n".join(parts)

    # 找講義，翰林國一數學
    match_material = re.match(r"找講義[，,]\s*(.+)", text)
    if match_material:
        keywords_str = match_material.group(1).strip()
        # 把關鍵字拆開（空格分隔，或整串搜）
        keywords = keywords_str.split() if " " in keywords_str else [keywords_str]
        files = search_materials(keywords)
        if not files:
            return f"找不到「{keywords_str}」相關的講義，請確認版本和書名。"
        lines = ["📚 找到以下講義：\n"]
        for name, link in files:
            lines.append(f"• {name}\n  {link}")
        return "\n".join(lines)

    # 說明指令
    if re.search(r"(幫助|help|指令|怎麼用)", text, re.IGNORECASE):
        return (
            "📖 使用方式：\n\n"
            "查上課時間：\n「本週上課時間，姓名」\n"
            "例：本週上課時間，王小明\n"
            "例：本週上課時間，劉冠毅\n\n"
            "找講義：\n「找講義，版本書名」\n"
            "例：找講義，翰林國一數學"
        )

    return None  # 不認識的訊息不回覆


@app.route("/webhook", methods=["POST"])
def webhook():
    signature = request.headers["X-Line-Signature"]
    body = request.get_data(as_text=True)

    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        abort(400)

    return "OK"


@handler.add(MessageEvent, message=TextMessage)
def on_message(event):
    reply = handle_message(event.message.text)
    if reply:
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text=reply)
        )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
